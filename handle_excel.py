import shutil
import subprocess
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


def run_command(command: str) -> Optional[str]:
    result = subprocess.run(
        ['powershell', '-Command', command],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        shell=False,
    )
    if result.returncode == 0:
        return result.stdout.strip()
    else:
        print(f'run command error: {command}')
        return None


def get_sys_model() -> str:
    sys_model = run_command('(Get-WmiObject -Class win32_computersystem).model')
    return '' if sys_model is None else sys_model


def find_rosa_fw_folders(base_dir: Path) -> Optional[Path]:
    for child in base_dir.iterdir():
        if 'rosa' in child.name.lower() and 'fw' in child.name.lower():
            return child
    return None


def find_files(directory: Path, pattern1: str, pattern2: str) -> Optional[Path]:
    for path in directory.rglob('*'):
        if pattern1 in path.name.lower() and pattern2 in path.name.lower():
            return path
    return None


def pull_fw_file(
    share_path: str, drive_letter: str, username: str, password: str
) -> Optional[Path]:
    command_del_letter = f'net use {drive_letter}: /delete'
    run_command(command_del_letter)
    command_map = f'net use {drive_letter}: {share_path} {password} /user:{username}'
    if run_command(command_map) is None:
        return None
    folder = find_rosa_fw_folders(Path(rf'{drive_letter}:/'))
    if folder is None:
        print(str(folder))
        return None
    excel_file = find_files(folder, 'key', 'device')
    if excel_file is None:
        return None
    file = shutil.copy2(excel_file, './')
    return file


# 两个字符串的最长公共子串
def longest_common_substring(s1: str, s2: str) -> str:
    if not isinstance(s1, str) or not isinstance(s2, str):
        raise ValueError(f'Both inputs must be strings, but got {type(s1)} and {type(s2)}')
    s1 = s1.lower()
    s2 = s2.lower()
    m = len(s1)
    n = len(s2)
    # 使用二维数组来存储最长连续公共子串的长度
    dp = [[0] * (n + 1) for _ in range(m + 1)]
    max_len = 0
    end = 0
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            if s1[i - 1] == s2[j - 1]:
                dp[i][j] = dp[i - 1][j - 1] + 1
                if dp[i][j] > max_len:
                    max_len = dp[i][j]
                    end = i  # 更新 end 位置
            else:
                dp[i][j] = 0

    if max_len > 0:
        start = end - max_len
        result = s1[start:end]
    else:
        result = ''

    return result


def get_selected_index(max_len: int) -> int:
    while True:
        _input = input('请输入1 or 2 or 3... 去选择你的project: ')
        try:
            selected_index = int(_input)
        except ValueError:
            print(f'亲, 请输入数字number({_input})')
            continue
        if selected_index > max_len:
            print(f'亲, 你输入{_input}这个没有')
            continue
        return selected_index


def get_inputed_project(projects: list[str]) -> str:
    projects = list(map(str.lower, projects))
    while True:
        _input = input('请输入你的project: ').lower()
        if _input not in projects:
            print(f'亲, 你确认有这个project-{_input}')
            continue
        return _input


def find_project_name(
    wb: Workbook, sheet_names: list[str], model_name: str, title_row: int
) -> Workbook:
    # device_with_sheet: Dict[str, List[str]] = dict()
    # print(sheet_names, '---', model_name)
    for sheet_name in sheet_names:
        all_project_name: List[Tuple[int, str]] = []
        if (
            sheet_name.startswith('ModelName')
            or sheet_name.startswith('Histroy')
            or sheet_name.startswith('Tool')
            or sheet_name.startswith('Cable')
        ):
            continue
        sheet = wb[sheet_name]

        try:
            row_2 = sheet[2]
        except IndexError:
            raise IndexError(f'{sheet}表没有第二行')
        index = 1
        cur_col = 1
        for cell in row_2:
            if sheet_name.startswith('Adapter'):
                font = Font(name='Calibri', size=14, bold=True, italic=False)
                cell.font = font
            if cell.value is None:
                continue
            cell_value = str(cell.value)
            # print(cell_value)
            # print(longest_common_substring(model_name, cell_value.lower()))
            # print(cell_value, '____', sheet_name)
            lcs = longest_common_substring(model_name, cell_value.lower())
            if len(lcs) > 4:
                all_project_name.append((cur_col, f'{index}: {cell_value}'))
                index += 1
            cur_col += 1
        if not all_project_name:
            sheet.insert_cols(8)
            # sheet.column_dimensions['H'].hidden = False
            sheet.cell(row=title_row, column=8).value = model_name
            sheet.cell(row=sheet.max_row + 1, column=8).value = 'V'
            sheet.cell(row=sheet.max_row, column=1, value=sheet[f'A{sheet.max_row-1}'].value)
            continue
        print(sheet_name, '-', [f'{name}' for _, name in all_project_name])
        selected_index = get_selected_index(len(all_project_name))
        if selected_index < 1 or selected_index > len(all_project_name):
            raise IndexError('选择的索引超出范围')
        selected_project = all_project_name[selected_index - 1]
        sheet_index, sel = selected_project
        # print(selected_project)
        cell = sheet.cell(row=title_row, column=sheet_index, value=model_name)
        # if sheet_name == 'Battery':
        #     break
        # device_with_sheet.setdefault(sheet_name, list()).extend(all_project_name)
    # print(device_with_sheet)
    copied_sheet = wb.copy_worksheet(sheet)
    wb.remove(sheet)
    copied_sheet.title = 'Adapter'
    wb.move_sheet('Adapter', offset=-2)
    print(wb.sheetnames)
    return wb


def env(is_dev: bool) -> str:
    global dev
    dev = r'\\172.16.2.2\Users\JinzhongLi'
    global release
    release = r'\\172.16.2.2\Users\"Harris Xu"'
    global projects
    projects = [
        'Jedi',
        'WASP',
        'Selek15',
        'Pinehills',
        'Red',
        'Hawk',
        'Bandon',
        'Northbay',
        'Selek G5',
        'Mockingbird',
        'Hellcat',
        'Shuri',
        'Moonknight',
        'Watchmen',
        'SOUTH PEAK',
        'Broadmoor',
        'Antman',
        'Cyborg',
        'Millennio',
        'Alienware',
        'Odin',
        'Stradale',
        'Odin',
        'Infinity',
        'Scorpio',
        'Arches',
        'Oasis',
        'Quake',
        'Sentry',
        'POLARIS',
    ]
    return dev if is_dev else release


def main():
    # 使用示例
    share_path = env(is_dev=False)
    local_path = 'X'
    file = pull_fw_file(share_path, local_path, 'User1', 'Us111111')

    if file is not None:
        print(share_path + str(file))

    # 加载现有的Excel文件
    if file is None:
        print('excel is null')
        exit(1)
    # file = './Rosa Key Device FW control_2024-10-28.xlsx'  # test file

    project = get_inputed_project(projects=projects)
    sys_model = get_sys_model()
    wb = load_workbook(filename=file, data_only=True)
    sheet_names = wb.sheetnames
    print(sheet_names)
    sheet_modelname = wb['ModelName']
    sheet_modelname['A3'] = project
    sheet_modelname['B3'] = sys_model  # type: ignore
    # sheet_modelname['B2'] = 'test'  # test file
    Modified_wb = find_project_name(wb, sheet_names, model_name=project, title_row=2)
    Modified_wb.save(filename='./Key_Device_FW_control.xlsx')
    Modified_wb.close()
    # wb.save(filename='Key_Device_FW_control.xlsx')


# print(sheet_modelname['B2'].value)


if __name__ == '__main__':
    # if not ctypes.windll.shell32.IsUserAnAdmin():
    #     ctypes.windll.shell32.ShellExecuteW(
    #         None, "runas", sys.executable, __file__, None, 1
    #     )
    # else:
    main()
